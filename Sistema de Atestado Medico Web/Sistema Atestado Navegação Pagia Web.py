import flet as ft
import openpyxl
from datetime import datetime, timedelta
import os
import shutil
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import io
import base64

# Aplicar estilo seaborn para gráficos mais suaves
sns.set(style="white")


def main(page: ft.Page):
    page.title = "Sistema de Atestados Médicos"
    page.theme_mode = ft.ThemeMode.LIGHT

    # Variáveis globais
    attached_files = []
    file_picker = ft.FilePicker()
    attach_picker = ft.FilePicker()
    page.overlay.extend([file_picker, attach_picker])

    # Função para ler dados do CID
    def read_cid_data(cid_code):
        try:
            wb = openpyxl.load_workbook(
                r"C:\Users\Eloizo\Desktop\Base sistema atestados\Tabela_Cid.xlsx"
            )
            sheet = wb["CIDS"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == cid_code:
                    return row[1]
            return "CID não encontrado"
        except Exception as e:
            print(f"Erro ao ler o arquivo CID: {e}")
            return "Erro ao ler o arquivo CID"

    # Função para abrir arquivo Excel e preencher campos
    def open_excel_file(path, matricula, fields, atestado_data_table):
        if not path or not matricula:
            return 0, 0
        if not os.path.exists(path):
            raise FileNotFoundError(f"O arquivo Excel não foi encontrado: {path}")
        wb = openpyxl.load_workbook(path)
        sheet = wb["Base"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == matricula:
                fields["nome"].value = row[1]
                fields["cargo"].value = row[2]
                fields["setor"].value = row[3]
                fields["responsavel"].value = row[4]
                fields["gestor"].value = row[5]
                fields["admissao"].value = (
                    row[6].strftime("%d/%m/%Y")
                    if isinstance(row[6], datetime)
                    else row[6]
                )
                fields["demissao"].value = (
                    row[7].strftime("%d/%m/%Y")
                    if isinstance(row[7], datetime)
                    else row[7]
                )
                if isinstance(row[6], datetime):
                    admissao_date = row[6]
                    today = datetime.today()
                    delta = today - admissao_date
                    years, months, days = (
                        delta.days // 365,
                        (delta.days % 365) // 30,
                        (delta.days % 365) % 30,
                    )
                    fields["tempo_casa"].value = (
                        f"{years} Anos {months} Meses {days} Dias"
                    )
                else:
                    fields["tempo_casa"].value = ""
                break
        return get_atestado_data(path, matricula, atestado_data_table)

    # Função para obter dados de atestados
    def get_atestado_data(path, matricula, atestado_data_table):
        wb = openpyxl.load_workbook(path)
        sheet = wb["Atestados"]
        atestado_data_table.rows.clear()
        total_atestados = 0
        total_dias = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[1]) == matricula:
                data_atestado = row[15]
                data_fim_atestado = row[16]
                dias_atestado = row[17]
                data_atestado_str = (
                    data_atestado.strftime("%d/%m/%Y")
                    if isinstance(data_atestado, datetime)
                    else str(data_atestado)
                )
                data_fim_atestado_str = (
                    data_fim_atestado.strftime("%d/%m/%Y")
                    if isinstance(data_fim_atestado, datetime)
                    else str(data_fim_atestado)
                )
                cid = row[12]
                cid_nome = row[13]
                instituicao = row[14]
                atestado_data_table.rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(data_atestado_str)),
                            ft.DataCell(ft.Text(data_fim_atestado_str)),
                            ft.DataCell(ft.Text(str(dias_atestado))),
                            ft.DataCell(ft.Text(cid)),
                            ft.DataCell(ft.Text(cid_nome)),
                            ft.DataCell(ft.Text(instituicao)),
                        ],
                        color=(
                            ft.colors.WHITE
                            if total_atestados % 2 == 0
                            else ft.colors.GREY_100
                        ),
                    )
                )
                total_atestados += 1
                total_dias += int(dias_atestado) if dias_atestado else 0
        wb.close()
        return total_atestados, total_dias

    # Função chamada quando o campo de matrícula perde o foco
    def on_matricula_blur(
        e,
        path_field,
        matricula_field,
        fields,
        atestado_data_table,
        total_atestados_text,
        total_dias_text,
    ):
        try:
            total_atestados, total_dias = open_excel_file(
                path_field.value, matricula_field.value, fields, atestado_data_table
            )
            total_atestados_text.value = f"Quantidade de atestados: {total_atestados}"
            total_dias_text.value = f"Total de dias: {total_dias}"
        except FileNotFoundError as fnf_error:
            print(f"Erro ao abrir arquivo: {fnf_error}")
            show_alert(e.page, f"Erro: {fnf_error}")
        except Exception as error:
            print(f"Erro inesperado: {error}")
            show_alert(e.page, f"Erro inesperado: {error}")
        e.page.update()

    # Função chamada quando o campo de dias de atestado perde o foco
    def on_dias_atestado_blur(
        e, data_atestado_field, dias_atestado_field, data_fim_atestado_field
    ):
        if data_atestado_field.value and dias_atestado_field.value.isdigit():
            dias = int(dias_atestado_field.value) - 1
            data_atestado = datetime.strptime(data_atestado_field.value, "%d/%m/%Y")
            data_fim = data_atestado + timedelta(days=dias)
            data_fim_atestado_field.value = data_fim.strftime("%d/%m/%Y")
            data_fim_atestado_field.update()

    # Função chamada quando o campo CID é alterado
    def on_cid_change(e, cid_field, cid_nome_field):
        cid_code = cid_field.value
        if cid_code:
            cid_description = read_cid_data(cid_code)
            cid_nome_field.value = cid_description
            cid_nome_field.update()

    # Função para exibir alertas
    def show_alert(page, message):
        alert_dialog = ft.AlertDialog(
            title=ft.Text("Atenção"),
            content=ft.Text(message),
            actions=[
                ft.TextButton(
                    "OK",
                    on_click=lambda e: (
                        setattr(alert_dialog, "open", False),
                        page.update(),
                    ),
                ),
            ],
        )
        page.dialog = alert_dialog
        alert_dialog.open = True
        page.update()

    # Função para limpar campos
    def clear_fields(fields, atestado_fields):
        for field in fields.values():
            field.value = ""
            field.update()
        for atestado_field in atestado_fields.values():
            atestado_field.value = ""
            atestado_field.update()

    # Função chamada quando um arquivo é selecionado
    def pick_file_result(e: ft.FilePickerResultEvent):
        if e.files:
            file_path_field.value = e.files[0].path
            page.update()

    # Função chamada quando arquivos são anexados
    def pick_files_result(e: ft.FilePickerResultEvent):
        if e.files:
            attached_files.extend(e.files)
            update_attached_files_text(page, attached_files)
        attach_picker.update()

    # Função para salvar arquivos anexados
    def save_attached_files(matricula, attached_files):
        base_folder = r"C:\Users\Eloizo\Desktop\Base sistema atestados\DadosArquivos"
        employee_folder = os.path.join(base_folder, matricula)
        if not os.path.exists(employee_folder):
            os.makedirs(employee_folder)
        saved_files = []
        for file in attached_files:
            file_extension = os.path.splitext(file.name)[1]
            new_filename = (
                f"{matricula}_{datetime.now().strftime('%Y%m%d%H%M%S')}{file_extension}"
            )
            destination = os.path.join(employee_folder, new_filename)
            shutil.copy(file.path, destination)
            saved_files.append(destination)
        return saved_files

    # Função para salvar atestado
    def save_atestado(e):
        if not matricula_field.value:
            show_alert(e.page, "Por favor, preencha a matrícula.")
            return
        file_path = file_path_field.value
        if not file_path:
            show_alert(e.page, "Por favor, selecione o arquivo Excel.")
            return
        wb = openpyxl.load_workbook(file_path)
        sheet = wb["Atestados"]
        last_row = sheet.max_row + 1
        sheet.cell(row=last_row, column=1, value=last_row - 1)
        sheet.cell(row=last_row, column=2, value=matricula_field.value)
        sheet.cell(row=last_row, column=3, value=fields["nome"].value)
        sheet.cell(row=last_row, column=4, value=fields["cargo"].value)
        sheet.cell(row=last_row, column=5, value=fields["setor"].value)
        sheet.cell(row=last_row, column=6, value=fields["responsavel"].value)
        sheet.cell(row=last_row, column=7, value=fields["gestor"].value)
        sheet.cell(row=last_row, column=8, value=fields["admissao"].value)
        sheet.cell(row=last_row, column=9, value=fields["demissao"].value)
        sheet.cell(row=last_row, column=10, value=fields["tempo_casa"].value)
        sheet.cell(row=last_row, column=11, value=atestado_fields["crm"].value)
        sheet.cell(row=last_row, column=12, value=atestado_fields["medico"].value)
        sheet.cell(row=last_row, column=13, value=atestado_fields["cid"].value)
        sheet.cell(row=last_row, column=14, value=atestado_fields["cid_nome"].value)
        sheet.cell(row=last_row, column=15, value=atestado_fields["instituicao"].value)
        sheet.cell(
            row=last_row, column=16, value=atestado_fields["data_atestado"].value
        )
        sheet.cell(
            row=last_row, column=17, value=atestado_fields["data_fim_atestado"].value
        )
        sheet.cell(
            row=last_row, column=18, value=atestado_fields["dias_atestado"].value
        )
        saved_files = []
        if attached_files:
            saved_files = save_attached_files(matricula_field.value, attached_files)
            sheet.cell(row=last_row, column=19, value=", ".join(saved_files))
        wb.save(file_path)
        wb.close()
        show_alert(
            e.page, f"Atestado salvo com sucesso! Arquivos anexados: {len(saved_files)}"
        )
        clear_fields(fields, atestado_fields)
        attached_files.clear()
        update_attached_files_text(e.page, attached_files)

    # Função para atualizar o texto de arquivos anexados
    def update_attached_files_text(page, attached_files):
        attached_files_text.value = f"Arquivos anexados: {len(attached_files)}"
        page.update()

    # Função para carregar dados e gerar gráficos
    def carregar_dados(e=None):
        try:
            base_folder = r"C:\Users\Eloizo\Desktop\Base sistema atestados"
            caminho = os.path.join(base_folder, "Dados.xlsx")

            if not os.path.exists(caminho):
                raise FileNotFoundError(
                    f"O arquivo 'Dados.xlsx' não foi encontrado em {base_folder}"
                )

            df = pd.read_excel(caminho, sheet_name="Atestados")

            total_dias = df["Dia"].sum()
            quantidade_matriculas = df["Matricula"].count()

            hoje = datetime.now()
            periodo_60_dias = hoje - timedelta(days=60)
            df["Data Inicio"] = pd.to_datetime(df["Data Inicio"])
            df_periodo = df[df["Data Inicio"] >= periodo_60_dias]
            total_dias_60 = df_periodo["Dia"].sum()

            card_total_dias.content = ft.Column(
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                controls=[
                    ft.Text(
                        "Total de dias", size=14, weight=ft.FontWeight.BOLD, max_lines=2
                    ),
                    ft.Text(f"{total_dias}", size=20),
                ],
            )

            card_quantidade.content = ft.Column(
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                controls=[
                    ft.Text(
                        "Quantidade:", size=14, weight=ft.FontWeight.BOLD, max_lines=2
                    ),
                    ft.Text(f"{quantidade_matriculas}", size=20),
                ],
            )

            card_dias_60.content = ft.Column(
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                controls=[
                    ft.Text(
                        "Total de atestado dentro do período de 60 dias",
                        size=14,
                        weight=ft.FontWeight.BOLD,
                        max_lines=2,
                    ),
                    ft.Text(f"{total_dias_60}", size=20),
                ],
            )

            fig_barra = gerar_grafico_barras(df)
            image_barra.src_base64 = fig_barra

            fig_linhas = gerar_grafico_linhas(df)
            image_linhas.src_base64 = fig_linhas

            page.update()

        except FileNotFoundError as e:
            print(f"Erro ao carregar planilha: {e}")
            show_alert(page, f"Erro: {e}")
        except Exception as e:
            print(f"Erro inesperado ao carregar planilha:  {e}")
            show_alert(page, f"Erro inesperado ao carregar dados: {e}")

    # Função para gerar gráfico de barras
    def gerar_grafico_barras(df):
        df_agrupado = df.groupby("Setor")["Dia"].sum().sort_values(ascending=True)
        cores = sns.light_palette("blue", as_cmap=False, n_colors=len(df_agrupado))
        plt.figure(figsize=(6, 4))
        ax = df_agrupado.plot(kind="barh", color=cores, edgecolor="none")
        plt.title("Total de dias por Setor", fontsize=14, weight="bold")
        plt.xlabel("Total de dias", fontsize=12)
        plt.ylabel("Setor", fontsize=12)

        plt.xticks(fontsize=10)
        plt.yticks(fontsize=10)

        plt.gca().patch.set_alpha(0)
        plt.gcf().set_facecolor("none")
        for spine in plt.gca().spines.values():
            spine.set_visible(False)
        for i, v in enumerate(df_agrupado):
            ax.text(v, i, f"{v:.0f}", ha="left", va="center")
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format="png", bbox_inches="tight", transparent=True)
        buf.seek(0)
        plt.close()
        return base64.b64encode(buf.read()).decode("utf-8")

    # Função para gerar gráfico de linhas
    def gerar_grafico_linhas(df):
        df["Mes"] = df["Data Inicio"].dt.strftime("%b")
        df_agrupado = df.groupby("Mes")["Dia"].sum()
        meses_ordenados = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
        df_agrupado = df_agrupado.reindex(meses_ordenados, fill_value=0)
        plt.figure(figsize=(6, 4))
        cores = sns.color_palette("coolwarm", 12)
        ax = df_agrupado.plot(
            kind="line",
            marker="o",
            color=cores[3],
            linewidth=2,
            markersize=8,
            markerfacecolor=cores[2],
        )
        plt.title("Total de dias por mês", fontsize=14, weight="bold")
        plt.xlabel("Mês", fontsize=12)
        plt.ylabel("Total de dias", fontsize=12)
        plt.xticks(fontsize=10)
        plt.yticks(fontsize=10)
        plt.gca().patch.set_alpha(0)
        plt.gcf().set_facecolor("none")
        for spine in plt.gca().spines.values():
            spine.set_visible(False)
        for i, v in enumerate(df_agrupado):
            ax.text(i, v, f"{v:.0f}", ha="center", va="bottom")
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format="png", bbox_inches="tight", transparent=True)
        buf.seek(0)
        plt.close()
        return base64.b64encode(buf.read()).decode("utf-8")

    # Componentes da interface
    file_path_field = ft.TextField(label="Caminho da Planilha Excel", width=720)
    matricula_field = ft.TextField(
        label="Matrícula",
        width=250,
        on_blur=lambda e: on_matricula_blur(
            e,
            file_path_field,
            matricula_field,
            fields,
            atestado_data_table,
            total_atestados_text,
            total_dias_text,
        ),
    )

    search_button = ft.IconButton(
        icon=ft.icons.SEARCH,
        on_click=lambda _: file_picker.pick_files(allowed_extensions=["xlsx"]),
    )

    # Função para anexar arquivos
    def attach_files(e):
        attach_picker.pick_files(
            allow_multiple=True, allowed_extensions=["pdf", "png", "jpg", "jpeg"]
        )

    # Botões de salvar e anexar arquivo
    save_button = ft.ElevatedButton(
        text="Salvar",
        on_click=save_atestado,
    )

    attach_button = ft.ElevatedButton(
        "Anexar Arquivo", icon=ft.icons.ATTACH_FILE, on_click=attach_files
    )

    attached_files_text = ft.Text("Arquivos anexados: 0")

    atestado_data_table = ft.DataTable(
        columns=[
            ft.DataColumn(
                ft.Text(
                    "Data Inicio",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "Data Fim",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "Dias",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "CID",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "CID Descrição",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "Instituição",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
        ],
        rows=[],
        bgcolor=ft.colors.BLUE_GREY_100,
    )

    scrollable_table = ft.Column(
        controls=[atestado_data_table],
        height=350,
        scroll="auto",
    )

    total_atestados_text = ft.Text("Quantidade de atestados: 0", weight="bold")
    total_dias_text = ft.Text("Total de dias: 0", weight="bold")

    atestado_fields = {
        "crm": ft.TextField(label="CRM Médico", width=250),
        "medico": ft.TextField(label="Nome Médico", width=250),
        "cid": ft.TextField(
            label="CID",
            width=100,
            on_change=lambda e: on_cid_change(
                e, atestado_fields["cid"], atestado_fields["cid_nome"]
            ),
        ),
        "cid_nome": ft.TextField(label="CID Nome", width=400, disabled=True),
        "instituicao": ft.TextField(label="Instituição", width=400),
        "data_atestado": ft.TextField(label="Data do Atestado", width=250),
        "dias_atestado": ft.TextField(
            label="Dias Atestado",
            width=100,
            on_blur=lambda e: on_dias_atestado_blur(
                e,
                atestado_fields["data_atestado"],
                atestado_fields["dias_atestado"],
                atestado_fields["data_fim_atestado"],
            ),
        ),
        "data_fim_atestado": ft.TextField(
            label="Data Fim Atestado", width=250, disabled=True
        ),
    }

    fields = {
        "nome": ft.TextField(label="Nome", width=250, disabled=True),
        "cargo": ft.TextField(label="Cargo", width=250, disabled=True),
        "setor": ft.TextField(label="Setor", width=250, disabled=True),
        "responsavel": ft.TextField(label="Responsável", width=250, disabled=True),
        "gestor": ft.TextField(label="Gestor", width=250, disabled=True),
        "admissao": ft.TextField(label="Admissão", width=250, disabled=True),
        "demissao": ft.TextField(label="Demissão", width=250, disabled=True),
        "tempo_casa": ft.TextField(label="Tempo de Casa", width=250, disabled=True),
    }

    # Componentes para análise de dados
    card_total_dias = ft.Card(width=200, height=100)
    card_quantidade = ft.Card(width=200, height=100)
    card_dias_60 = ft.Card(width=200, height=100)
    image_barra = ft.Image(width=500, height=400)
    image_linhas = ft.Image(width=600, height=400)

    # Função para voltar à tela inicial
    def go_home(e):
        page.go("/")

    # Função para mudança de rota
    def route_change(route):
        page.views.clear()
        page.views.append(
            ft.View(
                "/",
                [
                    appbar,
                    ft.Container(
                        content=ft.Text(
                            "Bem-vindo ao Sistema de Atestados Médicos",
                            size=20,
                            weight="bold",
                        ),
                        alignment=ft.alignment.center,
                    ),
                ],
            )
        )
        if page.route == "/dashboard":
            page.views.append(
                ft.View(
                    "/dashboard",
                    [
                        appbar,
                        ft.Container(
                            content=ft.Text(
                                "Análise de Atestados Médicos", size=18, weight="bold"
                            ),
                            alignment=ft.alignment.center,
                        ),
                        ft.Row(
                            [card_total_dias, card_quantidade, card_dias_60],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        ft.Row(
                            [image_barra, image_linhas],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        # Botão para voltar à tela inicial
                        ft.ElevatedButton(
                            text="Voltar para Tela Inicial", on_click=go_home
                        ),
                    ],
                )
            )
            carregar_dados()  # Carrega os dados automaticamente
        elif page.route == "/cadastro":
            page.views.append(
                ft.View(
                    "/cadastro",
                    [
                        appbar,
                        ft.Container(
                            content=ft.Text(
                                "Cadastro de Atestados", size=20, weight="bold"
                            ),
                            alignment=ft.alignment.center,
                        ),
                        ft.Row(
                            [file_path_field, search_button, matricula_field],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        ft.Row(
                            [
                                fields["nome"],
                                fields["cargo"],
                                fields["setor"],
                                fields["responsavel"],
                            ],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        ft.Row(
                            [
                                fields["gestor"],
                                fields["admissao"],
                                fields["demissao"],
                                fields["tempo_casa"],
                            ],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        ft.Container(
                            content=ft.Text(
                                "Cadastro novo atestado médico", size=14, weight="bold"
                            ),
                            alignment=ft.alignment.center,
                        ),
                        ft.Row(
                            [
                                atestado_fields["data_atestado"],
                                atestado_fields["dias_atestado"],
                                atestado_fields["data_fim_atestado"],
                                atestado_fields["instituicao"],
                            ],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        ft.Row(
                            [
                                atestado_fields["crm"],
                                atestado_fields["medico"],
                                atestado_fields["cid"],
                                atestado_fields["cid_nome"],
                            ],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        # Botões de Salvar e Anexar Arquivo lado a lado
                        ft.Row(
                            controls=[save_button, attach_button, attached_files_text],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        ft.Container(
                            content=ft.Text(
                                "Informações dos Atestados", size=14, weight="bold"
                            ),
                            alignment=ft.alignment.center,
                        ),
                        ft.Container(
                            content=scrollable_table, alignment=ft.alignment.center
                        ),
                        ft.Row(
                            [total_atestados_text, total_dias_text],
                            alignment=ft.MainAxisAlignment.CENTER,
                        ),
                        # Botão para voltar à tela inicial
                        ft.ElevatedButton(
                            text="Voltar para Tela Inicial", on_click=go_home
                        ),
                    ],
                )
            )
        page.update()

    def view_pop(view):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = route_change
    page.on_view_pop = view_pop

    appbar = ft.AppBar(
        leading=ft.Icon(ft.icons.HEALTH_AND_SAFETY),
        leading_width=40,
        title=ft.Text("Sistema de Atestados"),
        center_title=False,
        bgcolor=ft.colors.SURFACE_VARIANT,
        actions=[
            ft.IconButton(
                ft.icons.HOME, on_click=go_home
            ),  # Botão para voltar à tela inicial
            ft.IconButton(ft.icons.DASHBOARD, on_click=lambda _: page.go("/dashboard")),
            ft.IconButton(ft.icons.ADD_CIRCLE, on_click=lambda _: page.go("/cadastro")),
        ],
    )

    # Configuração dos eventos do file picker e attach picker
    file_picker.on_result = pick_file_result
    attach_picker.on_result = pick_files_result

    page.go(page.route)


ft.app(target=main, view=ft.AppView.WEB_BROWSER)
